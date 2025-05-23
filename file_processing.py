#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Functions for processing VK ID files and extracting phone numbers
"""

import asyncio
import os
import re
import time
import traceback
from datetime import datetime
from typing import List, Tuple, Optional, Union, Dict

import aiogram.utils.exceptions  # Добавляем этот импорт
import aiogram.utils.exceptions

from api_client import extract_phones_recursive
# Импорт функции advanced_search
from advanced_search import search_by_name_dob as advanced_search

# Импортируем openpyxl для работы с Excel
try:
    import openpyxl
    from openpyxl import Workbook

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("WARNING: openpyxl is not installed. Excel export will not work.")

from config import TEMP_DIR
from logger import logger


def extract_vk_links(text: str) -> List[str]:
    """
    Extract VK profile links from text

    Args:
        text (str): Text containing VK links

    Returns:
        List[str]: List of VK profile links
    """
    if not text:
        return []

    pattern = r'https?://(?:www\.)?vk\.com/id\d+'
    links = re.findall(pattern, text)
    return links


def extract_vk_id(link: str) -> Optional[str]:
    """
    Extract VK ID from a profile link

    Args:
        link (str): VK profile link

    Returns:
        Optional[str]: VK ID or None if not found
    """
    if not link:
        return None

    match = re.search(r'id(\d+)', link)
    if match:
        return match.group(1)
    return None


def extract_phone_from_vk_parsing(response: dict, target_vk_id: str) -> List[str]:
    """
    Extract phone numbers from API response, improved version with better pattern matching.

    Args:
        response (dict): API response
        target_vk_id (str): VK ID we are searching for

    Returns:
        List[str]: List of phone numbers or empty list if not found
    """
    if not response or not target_vk_id:
        return []

    if "error" in response or "List" not in response:
        return []

    logger.debug(f"Searching for phone numbers for VK ID: {target_vk_id}")

    try:
        # Используем улучшенную функцию извлечения телефонов
        phones = extract_phones_recursive(response)
        if phones:
            logger.info(f"Найдено {len(phones)} телефонов по улучшенному алгоритму для VK ID {target_vk_id}")
            return phones

        # Если не нашли телефоны, возвращаем пустой список
        logger.warning(f"No phone numbers found for VK ID {target_vk_id}")
        return []
    except Exception as e:
        logger.error(f"Error extracting phone numbers: {e}")
        logger.error(traceback.format_exc())
        return []

def extract_phones_batch(response: dict, vk_ids: List[str]) -> Dict[str, List[str]]:
    """
    Извлекает телефонные номера для нескольких VK ID из ответа API на пакетный запрос.

    Args:
        response (dict): Ответ API на пакетный запрос
        vk_ids (List[str]): Список VK ID, для которых был сделан запрос

    Returns:
        Dict[str, List[str]]: Словарь, где ключи - VK ID, значения - списки найденных телефонов
    """
    if not response or "error" in response or "List" not in response:
        return {vk_id: [] for vk_id in vk_ids}

    result = {vk_id: [] for vk_id in vk_ids}

    # Для каждой базы данных в ответе
    for db_name, db_info in response["List"].items():
        if not isinstance(db_info, dict) or "Data" not in db_info:
            continue

        data = db_info["Data"]
        if not data or not isinstance(data, list):
            continue

        # Создаем отображение VK ID -> индексы записей
        vk_id_to_indices = {}
        for i, record in enumerate(data):
            if not isinstance(record, dict):
                continue

            for field_name, field_value in record.items():
                if field_name.lower() in ["vkid", "vk_id", "vk id", "id", "userid"]:
                    str_value = str(field_value).strip()
                    if str_value in vk_ids:
                        if str_value not in vk_id_to_indices:
                            vk_id_to_indices[str_value] = []
                        vk_id_to_indices[str_value].append(i)

        # Поиск телефонов для каждого найденного VK ID
        for vk_id, indices in vk_id_to_indices.items():
            phones_for_id = []

            for idx in indices:
                # Проверяем текущую запись на наличие телефона
                current_record = data[idx]
                for field_name, field_value in current_record.items():
                    if any(phone_keyword in field_name.lower() for phone_keyword in
                           ["phone", "телефон", "тел", "tel", "мобильный"]) and field_value:
                        digits = ''.join(c for c in str(field_value) if c.isdigit())
                        if digits.startswith('79') and len(digits) >= 11:
                            if digits not in phones_for_id:
                                phones_for_id.append(digits)

                # Проверяем записи в окрестности ±3 записей
                for offset in range(-3, 4):
                    if offset == 0:  # Текущую запись уже проверили
                        continue

                    check_idx = idx + offset
                    if 0 <= check_idx < len(data):
                        check_record = data[check_idx]
                        if not isinstance(check_record, dict):
                            continue

                        for field_name, field_value in check_record.items():
                            if any(phone_keyword in field_name.lower() for phone_keyword in
                                   ["phone", "телефон", "тел", "tel", "мобильный"]) and field_value:
                                digits = ''.join(c for c in str(field_value) if c.isdigit())
                                if digits.startswith('79') and len(digits) >= 11:
                                    if digits not in phones_for_id:
                                        phones_for_id.append(digits)

            # Добавляем найденные телефоны в результат
            if phones_for_id:
                result[vk_id].extend(phones_for_id)

    # Удаляем дубликаты и сортируем
    for vk_id in result:
        result[vk_id] = sorted(list(set(result[vk_id])))

    return result

def create_results_file(results: List[Tuple[str, Union[List[str], None]]]) -> str:
    """
    Create an Excel file with the results, filtering for Russian mobile numbers only

    Args:
        results (List[Tuple[str, Union[List[str], None]]]): List of tuples (link, phones)

    Returns:
        str: Path to the created file
    """
    if not EXCEL_AVAILABLE:
        error_message = "openpyxl is not installed. Please install it using 'pip install openpyxl'"
        logger.error(error_message)
        raise ImportError(error_message)

    if not results:
        return ""

    # Create a timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"vk_results_{timestamp}.xlsx"
    file_path = TEMP_DIR / filename

    try:
        # Ensure directory exists
        os.makedirs(TEMP_DIR, exist_ok=True)

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "VK Results"

        # Add headers
        ws['A1'] = "VK Link"
        ws['B1'] = "Phone 1"
        ws['C1'] = "Phone 2"
        ws['D1'] = "Phone 3"
        ws['E1'] = "Phone 4"
        ws['F1'] = "Phone 5"

        # Fill data
        for row_idx, (link, phones) in enumerate(results, start=2):  # Start from row 2 (after header)
            # Always write the link in column A
            ws.cell(row=row_idx, column=1, value=link)

            if phones:
                # Фильтруем только российские мобильные номера (79...)
                russian_mobile = [phone for phone in phones if phone.startswith('79') and len(phone) >= 11]

                # Write each phone in a separate column
                for col_idx, phone in enumerate(russian_mobile, start=2):  # Start from column B
                    if col_idx <= 6:  # Limit to 5 phone numbers per link
                        ws.cell(row=row_idx, column=col_idx, value=phone)

        # Save the workbook
        wb.save(file_path)
        logger.info(f"Results file created: {file_path}")
        return str(file_path)
    except Exception as e:
        logger.error(f"Error creating results file: {e}")
        logger.error(traceback.format_exc())
        # Создадим простой TXT файл в случае ошибки с Excel
        try:
            txt_filename = f"vk_results_{timestamp}.txt"
            txt_file_path = TEMP_DIR / txt_filename

            with open(txt_file_path, 'w', encoding='utf-8') as f:
                f.write("VK Link\tPhone Numbers\n")
                for link, phones in results:
                    if phones:
                        russian_mobile = [phone for phone in phones if phone.startswith('79') and len(phone) >= 11]
                        f.write(f"{link}\t{', '.join(russian_mobile)}\n")
                    else:
                        f.write(f"{link}\tNo phones found\n")

            logger.info(f"Fallback TXT results file created: {txt_file_path}")
            return str(txt_file_path)
        except Exception as txt_error:
            logger.error(f"Failed to create fallback TXT file: {txt_error}")
            return ""


async def process_vk_links(items: List[str], user_id: int, chat_id: int, message_id: int, bot, process_vk_search, db,
                           is_name_dob_format: bool = False):
    """
    Process a list of VK links or name+DOB queries in batch mode with adaptive batch sizing

    Args:
        items (List[str]): List of VK profile links or name+dob queries
        user_id (int): User ID
        chat_id (int): Chat ID for progress updates
        message_id (int): Message ID for progress updates
        bot: Bot instance
        process_vk_search: Function to search VK IDs
        db: Database instance
        is_name_dob_format (bool): Set to True if items are in "LastName FirstName DD.MM.YYYY" format

    Returns:
        List[Tuple[str, List[str]]]: List of tuples (link/query, phones)
    """
    from api_client import api_client  # Импортируем клиент API

    if not items:
        return []

    results = []
    total = len(items)

    # Get user settings
    user_settings = db.get_user_settings(user_id)

    # Track successful and failed requests
    success_count = 0
    fail_count = 0

    # Флаг для отслеживания возможности редактирования сообщения
    can_edit_message = True
    update_message_id = message_id  # ID сообщения для обновления (может меняться, если создаем новое)

    # Тип данных для сообщений
    item_type = "запросов" if is_name_dob_format else "ссылок"

    # Добавляем небольшую задержку перед первым редактированием сообщения
    await asyncio.sleep(0.5)

    # Функция для безопасного обновления сообщения о прогрессе
    async def safe_update_progress(text):
        nonlocal can_edit_message, update_message_id

        if can_edit_message:
            try:
                await bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=update_message_id,
                    text=text
                )
            except (aiogram.utils.exceptions.MessageCantBeEdited,
                    aiogram.utils.exceptions.MessageNotModified,
                    aiogram.utils.exceptions.MessageToEditNotFound) as e:
                logger.warning(f"Cannot edit message: {e}. Will send new messages for progress updates.")
                can_edit_message = False
                # Отправляем новое сообщение вместо редактирования
                new_msg = await bot.send_message(
                    chat_id=chat_id,
                    text=text
                )
                update_message_id = new_msg.message_id
            except Exception as e:
                logger.error(f"Error updating progress message: {e}")
                can_edit_message = False
        else:
            # Если редактирование невозможно, отправляем новое сообщение
            try:
                new_msg = await bot.send_message(
                    chat_id=chat_id,
                    text=text
                )
                update_message_id = new_msg.message_id
            except Exception as e:
                logger.error(f"Error sending progress message: {e}")

    # Сообщаем о начале обработки
    await safe_update_progress(f"🔍 Подготовка к пакетной обработке {total} {item_type}...")

    # Разный подход для VK links и name+dob запросов
    if not is_name_dob_format:
        # ===================== Обработка VK ссылок =====================
        # Извлекаем VK IDs из ссылок
        vk_ids = []
        link_to_id_map = {}  # Отображение ссылка -> id для восстановления соответствия

        for link in items:
            vk_id = extract_vk_id(link)
            if vk_id:
                vk_ids.append(vk_id)
                link_to_id_map[vk_id] = link
            else:
                # Если не удалось извлечь ID, добавляем пустой результат
                results.append((link, []))
                fail_count += 1

        if not vk_ids:
            await safe_update_progress(f"❌ Не удалось извлечь ни одного валидного VK ID из {total} ссылок.")
            return results

        # Начинаем с меньшего размера батча и увеличиваем его при успешных запросах
        initial_batch_size = 1
        max_batch_size = 30
        current_batch_size = initial_batch_size

        # Разбиваем на пакеты начального размера
        batches = []
        for i in range(0, len(vk_ids), current_batch_size):
            batches.append(vk_ids[i:i + current_batch_size])

        # Обновляем статус
        await safe_update_progress(
            f"🔍 Найдено {len(vk_ids)} валидных ID. Разбито на {len(batches)} пакетов для обработки.\n"
            f"Начальный размер пакета: {current_batch_size} ID."
        )

        batch_results = {}  # Словарь для хранения результатов по ID

        # Счетчики успешных и неудачных пакетов для адаптации размера
        consecutive_success = 0
        consecutive_fails = 0

        # Обрабатываем каждый пакет
        batch_index = 0
        while batch_index < len(batches):
            batch = batches[batch_index]

            try:
                # Обновляем прогресс
                processed_ids = sum(len(b) for b in batches[:batch_index])
                progress_percent = processed_ids / len(vk_ids) * 100 if vk_ids else 0

                await safe_update_progress(
                    f"🔍 Обрабатываю пакет {batch_index + 1}/{len(batches)} ({len(batch)} ID)...\n"
                    f"Прогресс: {processed_ids}/{len(vk_ids)} ID ({progress_percent:.1f}%)\n"
                    f"✅ Успешно: {success_count} | ❌ Ошибок: {fail_count}"
                )

                # Делаем пакетный запрос с новым форматом (строка с разделителями \n)
                batch_response = await asyncio.to_thread(
                    api_client.search_batch,
                    batch,
                    lang=user_settings.get("language"),
                    limit=user_settings.get("limit", 2000)  # Используем больший лимит для пакетного запроса
                )

                # Обрабатываем результаты пакета
                if "error" in batch_response:
                    logger.error(f"Batch request error: {batch_response['error']}")

                    # Если ошибка содержит "таймаут" или код 500, уменьшаем размер пакета
                    if ("тайм-аут" in batch_response["error"].lower() or
                            "500" in batch_response["error"] or
                            "External server unavailable" in batch_response["error"]):
                        consecutive_fails += 1
                        consecutive_success = 0

                        # Уменьшаем размер текущего пакета и пересоздаем оставшиеся пакеты
                        if len(batch) > 3 and consecutive_fails >= 2:  # Не делаем пакеты слишком маленькими
                            new_batch_size = max(3, len(batch) // 2)  # Уменьшаем размер вдвое, но не меньше 3
                            logger.info(f"Reducing batch size to {new_batch_size} due to errors")

                            # Пересоздаем оставшиеся пакеты с новым размером
                            remaining_ids = []
                            for i in range(batch_index, len(batches)):
                                remaining_ids.extend(batches[i])

                            new_batches = []
                            for i in range(0, len(remaining_ids), new_batch_size):
                                new_batches.append(remaining_ids[i:i + new_batch_size])

                            # Обновляем список пакетов
                            batches = batches[:batch_index] + new_batches

                            await safe_update_progress(
                                f"⚠️ Уменьшаем размер пакета до {new_batch_size} ID из-за ошибок API.\n"
                                f"Новое количество пакетов: {len(batches)}"
                            )

                            # Повторяем текущий индекс (не инкрементируем batch_index)
                            continue

                    # При ошибке помечаем все ID в пакете как неудачные
                    for vk_id in batch:
                        batch_results[vk_id] = []
                        fail_count += 1
                else:
                    consecutive_success += 1
                    consecutive_fails = 0

                    # Используем улучшенную функцию извлечения телефонов
                    phones_by_id = extract_phones_improved(batch_response, batch)

                    # Обрабатываем результаты по каждому ID
                    for vk_id in batch:
                        vk_id_clean = vk_id.strip()  # Очищаем ID от пробелов
                        phones = phones_by_id.get(vk_id_clean, [])
                        batch_results[vk_id] = phones

                        if phones:
                            success_count += 1
                        else:
                            fail_count += 1

                    # Если несколько успешных запросов подряд, увеличиваем размер пакета для оставшихся
                    if consecutive_success >= 3 and current_batch_size < max_batch_size:
                        # Увеличиваем размер пакета, но не больше максимального
                        new_batch_size = min(max_batch_size, current_batch_size * 2)

                        if new_batch_size > current_batch_size:
                            current_batch_size = new_batch_size
                            logger.info(f"Increasing batch size to {current_batch_size} due to successful requests")

                            # Пересоздаем оставшиеся пакеты с новым размером
                            if batch_index + 1 < len(batches):
                                remaining_ids = []
                                for i in range(batch_index + 1, len(batches)):
                                    remaining_ids.extend(batches[i])

                                new_batches = []
                                for i in range(0, len(remaining_ids), current_batch_size):
                                    new_batches.append(remaining_ids[i:i + current_batch_size])

                                # Обновляем список пакетов
                                batches = batches[:batch_index + 1] + new_batches

                                await safe_update_progress(
                                    f"✅ Увеличиваем размер пакета до {current_batch_size} ID.\n"
                                    f"Новое количество пакетов: {len(batches)}"
                                )

                # Инкрементируем индекс пакета
                batch_index += 1

                # Небольшая задержка между пакетами для снижения нагрузки
                await asyncio.sleep(4)

            except Exception as e:
                logger.error(f"Error processing batch {batch_index + 1}: {e}")
                logger.error(traceback.format_exc())
                # При исключении помечаем все ID в пакете как неудачные
                for vk_id in batch:
                    batch_results[vk_id] = []
                    fail_count += 1

                # Инкрементируем индекс пакета и продолжаем
                batch_index += 1

        # Формируем итоговый результат, сохраняя порядок оригинальных ссылок
        for link in items:
            vk_id = extract_vk_id(link)
            if vk_id and vk_id in batch_results:
                results.append((link, batch_results[vk_id]))
            elif (link, []) not in results:  # Проверяем, не добавили ли мы уже эту ссылку ранее
                results.append((link, []))

    else:
        # ===================== Обработка запросов ФИО + дата рождения =====================
        # Прямой перебор в цикле (без пакетной обработки)
        for i, query in enumerate(items, 1):
            progress_percent = (i / total) * 100

            # Обновляем статус каждые 5 запросов или для первого/последнего
            if i == 1 or i == total or i % 5 == 0:
                await safe_update_progress(
                    f"🔍 Обработка запроса {i}/{total} ({progress_percent:.1f}%)...\n"
                    f"✅ Успешно: {success_count} | ❌ Ошибок: {fail_count}"
                )

            try:
                # Разбор запроса
                parts = query.split()
                if len(parts) < 3:
                    results.append((query, []))
                    fail_count += 1
                    continue

                # Проверяем, является ли последний элемент датой
                dob_pattern = re.compile(r'^\d{1,2}[./-]\d{1,2}[./-]\d{2,4}$')
                if not dob_pattern.match(parts[-1]):
                    results.append((query, []))
                    fail_count += 1
                    continue

                # Извлечение компонентов запроса
                date_of_birth = parts[-1]
                last_name = parts[0]
                first_name = parts[1] if len(parts) > 2 else ""

                # Этап 1: Поиск по ФИО + дате рождения
                response1 = await asyncio.to_thread(
                    api_client.make_request,
                    query=query,
                    lang=user_settings.get("language", "ru"),
                    limit=user_settings.get("limit", 2000)
                )

                # Анализ результатов и извлечение email, телефонов и VK ID
                emails, phones, vk_ids = [], [], []

                # Приоритетные источники
                source_priority = [
                    "Gosuslugi 2024", "BolshayaPeremena", "AlfaBank 2023 v2",
                    "Resh.Edu", "ProPostuplenie.ru", "TrudVsem.ru"
                ]

                if "List" in response1 and isinstance(response1["List"], dict):
                    # Сначала ищем в приоритетных источниках
                    for source_name in source_priority:
                        if source_name in response1["List"]:
                            source_data = response1["List"][source_name]
                            if "Data" in source_data and isinstance(source_data["Data"], list):
                                for record in source_data["Data"]:
                                    # Ищем телефоны и email
                                    for field_name, field_value in record.items():
                                        field_lower = field_name.lower()
                                        if "номер который нужно забирать" in str(record):
                                            # Если есть явная пометка, берем этот номер
                                            if "phone" in field_lower or "телефон" in field_lower:
                                                digits = ''.join(c for c in str(field_value) if c.isdigit())
                                                if digits.startswith('79') and len(
                                                        digits) == 11 and digits not in phones:
                                                    phones.append(digits)

                                        # Ищем телефон в обычном режиме
                                        if "phone" in field_lower or "телефон" in field_lower:
                                            digits = ''.join(c for c in str(field_value) if c.isdigit())
                                            if digits.startswith('79') and len(digits) == 11 and digits not in phones:
                                                phones.append(digits)

                                        # Ищем email
                                        if "email" in field_lower and "@" in str(field_value):
                                            email = str(field_value).strip().lower()
                                            if email not in emails:
                                                emails.append(email)

                                        # Ищем VK ID
                                        if "vk" in field_lower or "id" in field_lower:
                                            vk_match = re.search(r'(?:vk\.com\/id|\/id|^id)(\d+)', str(field_value))
                                            if vk_match and vk_match.group(1) not in vk_ids:
                                                vk_ids.append(vk_match.group(1))

                # Этап 2: Если нашли email, выполняем поиск по нему
                if not phones and emails:
                    response2 = await asyncio.to_thread(
                        api_client.make_request,
                        query=emails[0],
                        lang=user_settings.get("language", "ru"),
                        limit=user_settings.get("limit", 2000)
                    )

                    # Анализируем результаты поиска по email
                    if "List" in response2 and isinstance(response2["List"], dict):
                        for source_name in source_priority:
                            if source_name in response2["List"]:
                                source_data = response2["List"][source_name]
                                if "Data" in source_data and isinstance(source_data["Data"], list):
                                    for record in source_data["Data"]:
                                        # Ищем телефоны
                                        if "номер который нужно забирать" in str(record):
                                            # Если есть явная пометка, берем этот номер
                                            for field_name, field_value in record.items():
                                                field_lower = field_name.lower()
                                                if "phone" in field_lower or "телефон" in field_lower:
                                                    digits = ''.join(c for c in str(field_value) if c.isdigit())
                                                    if digits.startswith('79') and len(
                                                            digits) == 11 and digits not in phones:
                                                        phones.append(digits)

                                        # Ищем телефон обычным способом
                                        for field_name, field_value in record.items():
                                            field_lower = field_name.lower()
                                            if "phone" in field_lower or "телефон" in field_lower:
                                                digits = ''.join(c for c in str(field_value) if c.isdigit())
                                                if digits.startswith('79') and len(
                                                        digits) == 11 and digits not in phones:
                                                    phones.append(digits)

                # Этап 3: Если все еще нет результатов и найден VK ID, пробуем через него
                if not phones and vk_ids:
                    vk_response = await process_vk_search(vk_ids[0], user_id, user_settings)
                    vk_phones = extract_phone_from_vk_parsing(vk_response, vk_ids[0])
                    phones.extend([p for p in vk_phones if p not in phones])

                # Добавляем результаты в общий список
                results.append((query, phones))

                # Обновляем статистику
                if phones:
                    success_count += 1
                else:
                    fail_count += 1

                # Небольшая задержка между запросами
                await asyncio.sleep(3)

            except Exception as e:
                logger.error(f"Error processing name+dob query {i}/{total}: {e}")
                logger.error(traceback.format_exc())
                results.append((query, []))
                fail_count += 1

    # Обновляем финальный статус
    phones_found = sum(1 for _, phones in results if phones and len(phones) > 0)
    total_phones = sum(len(phones) for _, phones in results if phones)

    await safe_update_progress(
        f"✅ Обработка завершена: {len(items)} {item_type} (100%)\n"
        f"✅ Успешно: {success_count} | ❌ Ошибок: {fail_count}\n"
        f"📱 Найдено {phones_found} записей с номерами (всего {total_phones} номеров)"
    )

    return results


def extract_phones_improved(response: dict, batch_vk_ids: List[str]) -> Dict[str, List[str]]:
    """
    Улучшенная функция извлечения телефонов из ответа API на пакетный запрос.
    Учитывает особенности формата данных, включая пробелы в полях VkID.
    """
    if not response or "error" in response or "List" not in response:
        return {vk_id: [] for vk_id in batch_vk_ids}

    # Подготовка результатов с нормализацией ID (очистка пробелов)
    result = {vk_id.strip(): [] for vk_id in batch_vk_ids}
    clean_batch_ids = [vk_id.strip() for vk_id in batch_vk_ids]

    # Проверяем явное указание "Номер который нужно забирать"
    marked_phones = {}  # vk_id -> phone

    # Для каждой базы данных в ответе
    for db_name, db_info in response["List"].items():
        if not isinstance(db_info, dict) or "Data" not in db_info:
            continue

        data = db_info["Data"]
        if not data or not isinstance(data, list):
            continue

        # Приоритетные источники
        source_priority = {
            "Gosuslugi 2024": 15,
            "BolshayaPeremena": 14,
            "AlfaBank 2023 v2": 13,
            "ScanTour.ru": 12,
            "Resh.Edu": 11,
            "ProPostuplenie.ru": 10,
            "TrudVsem.ru": 9,
            "Dobro.ru": 8,
            "CDEK": 7,
            "Whoosh-bike": 6,
            "DNS (2022)": 5,
            "SushiMaster.ru": 5,
            "BurgerKing.ru": 5,
            "GloriaJeans": 5,
            "LeaderID": 4,
            "Book24": 4,
            "Rendez-Vous": 4,
            "Zoloto585.ru": 4,
            "Tokyo-city.ru": 3,
            "Pikabu": 3,
            "Metro-cc.ru": 3,
            "Adengi.ru": 3,
            "Mira1.ru": 3,
            "Orteka.ru": 2,
            "Oriflame.ru": 2,
            "Artek": 2
        }
        priority = source_priority.get(db_name, 0)

        # Обрабатываем каждую запись
        for record in data:
            if not isinstance(record, dict):
                continue

            # Проверяем на наличие пометки "Номер который нужно забирать"
            record_str = str(record)
            has_marker = "Номер который нужно забирать" in record_str

            # 1. Ищем VK ID в текущей записи
            found_vk_id = None
            for field_name, field_value in record.items():
                field_lower = field_name.lower()
                # Проверяем поля, которые могут содержать VK ID
                if "vkid" in field_lower or "vk_id" in field_lower or field_lower == "id":
                    # Обязательно очищаем значение от пробелов!
                    str_value = str(field_value).strip()
                    if str_value in clean_batch_ids:
                        found_vk_id = str_value
                        break

            # 2. Если нашли VK ID, ищем телефон
            if found_vk_id:
                for field_name, field_value in record.items():
                    field_lower = field_name.lower()
                    # Ищем поля с телефонами
                    if "phone" in field_lower or "телефон" in field_lower or "тел" in field_lower:
                        if field_value:
                            # Извлекаем только цифры
                            digits = ''.join(c for c in str(field_value) if c.isdigit())
                            # Проверяем формат российского мобильного (79XXXXXXXXX)
                            if digits.startswith('79') and len(digits) == 11:
                                # Если есть пометка "Номер который нужно забирать"
                                if has_marker:
                                    marked_phones[found_vk_id] = digits
                                # Иначе добавляем в обычный список
                                elif digits not in result[found_vk_id]:
                                    result[found_vk_id].append(digits)

    # Добавляем помеченные номера в начало списков
    for vk_id, phone in marked_phones.items():
        if phone in result[vk_id]:
            result[vk_id].remove(phone)
        result[vk_id].insert(0, phone)

    return result


async def process_name_dob_queries(queries, user_id, chat_id, message_id, bot_instance, db_instance):
    """
    Последовательная (не пакетная) обработка запросов ФИО + дата рождения

    Args:
        queries (List[str]): Список запросов в формате "Фамилия Имя ДД.ММ.ГГГГ"
        user_id (int): ID пользователя
        chat_id (int): ID чата
        message_id (int): ID сообщения для обновления
        bot_instance: Экземпляр бота
        db_instance: Экземпляр базы данных

    Returns:
        List[Tuple[str, List[str]]]: Список пар (запрос, список телефонов)
    """
    if not queries:
        return []

    results = []
    total = len(queries)
    processed = 0

    # Получаем настройки пользователя
    user_settings = db_instance.get_user_settings(user_id)

    # Счетчики для статистики
    success_count = 0
    fail_count = 0

    # Флаг для отслеживания возможности редактирования сообщения
    can_edit_message = True
    update_message_id = message_id

    # Функция для безопасного обновления сообщения о прогрессе
    async def safe_update_progress(text):
        nonlocal can_edit_message, update_message_id

        if can_edit_message:
            try:
                await bot_instance.edit_message_text(
                    chat_id=chat_id,
                    message_id=update_message_id,
                    text=text
                )
            except Exception as e:
                logger.warning(f"Не удалось отредактировать сообщение: {e}. Отправляем новое.")
                can_edit_message = False
                new_msg = await bot_instance.send_message(
                    chat_id=chat_id,
                    text=text
                )
                update_message_id = new_msg.message_id
        else:
            try:
                new_msg = await bot_instance.send_message(
                    chat_id=chat_id,
                    text=text
                )
                update_message_id = new_msg.message_id
            except Exception as e:
                logger.error(f"Ошибка при отправке сообщения о прогрессе: {e}")

    await safe_update_progress(
        f"🔍 Начинаю последовательную обработку {len(queries)} запросов.\n"
        f"⏳ Пожалуйста, подождите..."
    )

    # Обрабатываем каждый запрос по отдельности
    for i, query in enumerate(queries):
        # Обновляем прогресс каждые 3 запроса или для первого/последнего
        if i % 3 == 0 or i == 0 or i == total - 1:
            progress_percent = (i / total) * 100
            await safe_update_progress(
                f"🔄 Обработка запроса {i + 1}/{total} ({progress_percent:.1f}%)...\n"
                f"✅ Успешно: {success_count} | ❌ Ошибок: {fail_count}\n"
                f"🔍 Текущий запрос: {query}"
            )

        try:
            # Обрабатываем запрос и дожидаемся результата
            result = await process_single_name_dob_query(query, user_id, user_settings)

            # Извлекаем телефоны из результата
            phones = result.get("phones", [])

            # Добавляем результат в общий список
            results.append((query, phones))

            # Обновляем статистику
            if phones:
                success_count += 1
            else:
                fail_count += 1

            # Добавляем более длительную паузу между запросами
            await asyncio.sleep(2.0)  # Увеличенная пауза для гарантии полной обработки

        except Exception as e:
            logger.error(f"Ошибка при обработке запроса '{query}': {e}")
            logger.error(traceback.format_exc())
            results.append((query, []))
            fail_count += 1
            # Продолжаем с следующим запросом после ошибки

        processed += 1

    # Финальное обновление прогресса
    phones_found = sum(1 for _, phones in results if phones and len(phones) > 0)
    total_phones = sum(len(phones) for _, phones in results if phones)

    await safe_update_progress(
        f"✅ Обработка завершена: {total} запросов (100%)\n"
        f"✅ Успешно: {success_count} | ❌ Ошибок: {fail_count}\n"
        f"📱 Найдено {phones_found} записей с номерами (всего {total_phones} номеров)"
    )

    return results


async def process_single_name_dob_query(query, user_id, user_settings):
    """
    Обработка одного запроса ФИО + дата рождения

    Args:
        query (str): Запрос в формате "Фамилия Имя ДД.ММ.ГГГГ"
        user_id (int): ID пользователя
        user_settings (dict): Настройки пользователя

    Returns:
        dict: Результат поиска
    """
    try:
        # Преобразуем формат даты из ДД.ММ.ГГГГ в ГГГГ-ММ-ДД для API
        parts = query.split()
        name_parts = parts[:-1]  # Все кроме последней части (даты)
        date_part = parts[-1]  # Последняя часть - дата

        # Преобразуем дату, если она в формате ДД.ММ.ГГГГ
        if '.' in date_part:
            day, month, year = date_part.split('.')
            iso_date = f"{year}-{month}-{day}"
        else:
            iso_date = date_part

        # Формируем строку запроса с ISO-форматом даты
        search_query = f"{' '.join(name_parts)} {iso_date}"

        # Запускаем синхронную функцию в отдельном потоке
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(
            None,
            lambda: advanced_search(
                query=search_query,
                lang=user_settings.get("language", "ru"),
                limit=user_settings.get("limit", 300)
            )
        )

        # Проверяем и логируем результаты
        if "phones" in result and result["phones"]:
            logger.info(f"Найдено {len(result['phones'])} телефонов для запроса '{query}'")
            for phone in result["phones"]:
                logger.info(f"  - Телефон: {phone}")

        return result
    except Exception as e:
        logger.error(f"Ошибка при обработке запроса '{query}': {e}")
        logger.error(traceback.format_exc())
        return {
            "query": query,
            "phones": [],
            "method": "error",
            "confidence": 0.0,
            "error": str(e)
        }

def analyze_first_stage_results(response: dict, query: str) -> Tuple[List[str], List[str], float, List[str]]:
    """
    Анализирует результаты первого этапа поиска (по ФИО и дате рождения)
    с улучшенным извлечением email, телефонов и других идентификаторов.

    Args:
        response (dict): Ответ от API
        query (str): Исходный запрос "Фамилия Имя ДД.ММ.ГГГГ"

    Returns:
        Tuple[List[str], List[str], float, List[str]]:
            (emails, телефоны, уверенность, vk_ids)
    """
    # Логируем полный ответ API для отладки
    logger.debug(f"Анализ ответа API для запроса '{query}'")

    # Извлекаем компоненты из запроса
    parts = query.split()

    # Проверяем, есть ли дата рождения в запросе
    dob_pattern = re.compile(r'^\d{1,2}[./-]\d{1,2}[./-]\d{2,4}$')

    if len(parts) >= 3 and dob_pattern.match(parts[-1]):
        date_of_birth = parts[-1]
        name_parts = parts[:-1]

        # Предполагаем, что первое слово - фамилия, второе - имя
        last_name = name_parts[0].lower()
        first_name = name_parts[1].lower() if len(name_parts) > 1 else ""
    else:
        date_of_birth = None
        name_parts = parts
        last_name = name_parts[0].lower() if name_parts else ""
        first_name = name_parts[1].lower() if len(name_parts) > 1 else ""

    # Результаты анализа
    emails = []
    vk_ids = []

    # Используем улучшенную функцию для извлечения телефонов
    phones = extract_phones_from_api_response(response)

    # УЛУЧШЕНИЕ 1: Преобразуем весь ответ API в строку для текстового анализа
    response_str = str(response)

    # УЛУЧШЕНИЕ 2: Расширенный набор паттернов для поиска email
    email_patterns = [
        r'📩Email:\s*([^\s,]+@[^\s,]+)',
        r'Email:\s*([^\s,]+@[^\s,]+)',
        r'email:\s*([^\s,]+@[^\s,]+)',
        r'E-mail:\s*([^\s,]+@[^\s,]+)',
        r'e-mail:\s*([^\s,]+@[^\s,]+)',
        r'[\'"]email[\'"]:\s*[\'"]([^\'"]+@[^\'"]+)[\'"]',
        r'[\'"]Email[\'"]:\s*[\'"]([^\'"]+@[^\'"]+)[\'"]',
        r'[^a-zA-Z0-9]([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'  # общий шаблон для email
    ]

    for pattern in email_patterns:
        email_matches = re.findall(pattern, response_str, re.IGNORECASE)
        for email in email_matches:
            email = email.strip().lower()
            if '@' in email and email not in emails:
                emails.append(email)
                logger.info(f"Найден email по паттерну: {email}")

    # УЛУЧШЕНИЕ 3: Расширенный набор паттернов для поиска VK ID
    vk_patterns = [
        r'🆔VK ID:\s*([^\s]+)',
        r'VK ID:\s*([^\s]+)',
        r'vk\.com/id(\d+)',
        r'https://vk\.com/id(\d+)',
        r'[\'"]vk_id[\'"]:\s*[\'"]([^\'"]+)[\'"]',
        r'[\'"]VkID[\'"]:\s*[\'"]([^\'"]+)[\'"]',
        r'[\'"]vk_com[\'"]:\s*[\'"]([^\'"]+)[\'"]'
    ]

    for pattern in vk_patterns:
        vk_matches = re.findall(pattern, response_str, re.IGNORECASE)
        for vk_match in vk_matches:
            if vk_match and (vk_match.isdigit() or (
                    isinstance(vk_match, str) and vk_match.startswith('id') and vk_match[2:].isdigit())):
                # Нормализуем VK ID
                vk_id = vk_match[2:] if isinstance(vk_match, str) and vk_match.startswith('id') else vk_match
                if str(vk_id).isdigit() and vk_id not in vk_ids:
                    vk_ids.append(str(vk_id))
                    logger.info(f"Найден VK ID по паттерну: {vk_id}")

    # Определяем уровень уверенности
    max_confidence = 0.0
    if phones:
        max_confidence = 0.8  # Высокая уверенность, если нашли телефоны
    elif emails:
        max_confidence = 0.6  # Средняя уверенность для продолжения поиска
        logger.info(f"Не найдены телефоны, но найдены email: {emails}. Подготовка к второму этапу поиска.")

    # УЛУЧШЕНИЕ 4: Более глубокий анализ структуры ответа для поиска email
    if "List" in response and isinstance(response["List"], dict):
        # Анализируем каждую базу данных в ответе
        for source_name, source_data in response["List"].items():
            if source_name == "No results found":
                continue

            # Проверка наличия данных
            if "Data" in source_data and isinstance(source_data["Data"], list):
                for record in source_data["Data"]:
                    if not isinstance(record, dict):
                        continue

                    # УЛУЧШЕНИЕ 5: Проверка всех полей на наличие email
                    for field, value in record.items():
                        # Любые поля, которые могут содержать email
                        if isinstance(value, str) and '@' in value and '.' in value:
                            email_candidate = value.lower().strip()
                            # Проверка на валидность email
                            if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email_candidate):
                                if email_candidate not in emails:
                                    emails.append(email_candidate)
                                    logger.info(f"Найден email в поле {field}: {email_candidate}")

                        # Конкретные поля с email
                        if "email" in str(field).lower() and isinstance(value, str) and '@' in value:
                            email = value.lower().strip()
                            if email not in emails:
                                emails.append(email)
                                logger.info(f"Найден email в поле {field}: {email}")

    # УЛУЧШЕНИЕ 6: Оценка важности email на основе контекста запроса
    if emails:
        # Сортируем email по приоритету
        email_priority = []
        for email in emails:
            priority = 0

            # Высокий приоритет если email содержит фамилию или имя
            if last_name and last_name in email.split('@')[0]:
                priority += 10
            if first_name and first_name in email.split('@')[0]:
                priority += 5

            # Приоритет по популярным доменам
            domain = email.split('@')[-1].lower()
            if domain in ['gmail.com', 'mail.ru', 'yandex.ru', 'bk.ru', 'inbox.ru', 'list.ru', 'icloud.com']:
                priority += 3

            email_priority.append((email, priority))

        # Сортируем по приоритету (от высокого к низкому)
        email_priority.sort(key=lambda x: x[1], reverse=True)

        # Обновляем список email
        emails = [e[0] for e in email_priority]
        logger.info(f"Отсортированные email по приоритету: {emails}")

    # УЛУЧШЕНИЕ 7: Проверка наличия более специфической информации для увеличения уверенности
    if "List" in response and isinstance(response["List"], dict):
        # Ищем приоритетные базы данных
        priority_sources = ["Gosuslugi 2024", "BolshayaPeremena", "AlfaBank 2023 v2", "ScanTour.ru"]
        for source_name in priority_sources:
            if source_name in response["List"]:
                max_confidence = max(max_confidence, 0.7)  # Повышаем уверенность
                logger.info(f"Найдена приоритетная база данных {source_name}, уверенность повышена до {max_confidence}")

    logger.info(
        f"Результаты анализа первого этапа: {len(emails)} email, {len(phones)} телефонов, уверенность {max_confidence:.2f}")
    return emails, phones, max_confidence, vk_ids

def analyze_second_stage_results(response: dict, original_query: str, email: str) -> Tuple[List[str], float]:
    """
    Анализирует результаты второго этапа поиска (по email)

    Args:
        response (dict): Ответ от API
        original_query (str): Исходный запрос "Фамилия Имя ДД.ММ.ГГГГ"
        email (str): Email, по которому был сделан запрос

    Returns:
        Tuple[List[str], float]: (телефоны, уверенность)
    """
    # Логируем полный ответ API для отладки
    logger.debug(f"Email search API Response full content: {response}")

    # Извлекаем компоненты из запроса
    parts = original_query.split()

    # Проверяем, есть ли дата рождения в запросе
    dob_pattern = re.compile(r'^\d{1,2}[./-]\d{1,2}[./-]\d{2,4}$')

    if len(parts) >= 3 and dob_pattern.match(parts[-1]):
        date_of_birth = parts[-1]
        name_parts = parts[:-1]
        last_name = name_parts[0].lower()
        first_name = name_parts[1].lower() if len(name_parts) > 1 else ""
    else:
        date_of_birth = None
        name_parts = parts
        last_name = name_parts[0].lower() if name_parts else ""
        first_name = name_parts[1].lower() if len(name_parts) > 1 else ""

    # Результаты анализа
    phones = []
    max_confidence = 0.0

    # Преобразуем ответ API в строку для текстового анализа
    response_str = str(response)

    # ВАЖНО: Проверка на наличие пометок о телефонах
    if "Номер который нужно забирать" in response_str:
        logger.info(f"Найдена пометка 'Номер который нужно забирать' в ответе по email")

        # Несколько вариантов регулярных выражений для поиска
        patterns = [
            r'📞Телефон:\s*(\d+)[^)]*Номер который нужно забирать',
            r'телефон:\s*(\d+)[^)]*Номер который нужно забирать',
            r'телефон.*?:\s*(\d+)[^)]*Номер который нужно забирать',
            r'телефона?[^:]*:\s*(\d+)[^)]*Номер который нужно забирать'
        ]

        for pattern in patterns:
            marked_matches = re.findall(pattern, response_str, re.IGNORECASE)
            for phone in marked_matches:
                digits = ''.join(c for c in phone if c.isdigit())
                if digits.startswith('79') and len(digits) == 11 and digits not in phones:
                    phones.append(digits)
                    max_confidence = 0.95  # Высокая уверенность
                    logger.info(f"Найден телефон с пометкой в ответе по email: {digits}")

    # Поиск всех телефонов с эмодзи
    phone_patterns = [
        r'📞Телефон:\s*(\d+)',
        r'Телефон:\s*(\d+)',
        r'телефон:\s*(\d+)'
    ]

    for pattern in phone_patterns:
        phone_matches = re.findall(pattern, response_str, re.IGNORECASE)
        for phone in phone_matches:
            digits = ''.join(c for c in phone if c.isdigit())
            if digits.startswith('79') and len(digits) == 11 and digits not in phones:
                phones.append(digits)
                if max_confidence < 0.8:
                    max_confidence = 0.8  # Хорошая уверенность
                logger.info(f"Найден телефон в ответе по email: {digits}")

    # Проверяем совпадение фамилии и даты рождения для подтверждения
    if last_name and date_of_birth:
        if last_name.lower() in response_str.lower() and (
                date_of_birth in response_str or
                date_of_birth.replace('.', '-') in response_str or
                date_of_birth.replace('.', '/') in response_str
        ):
            # Повышаем уверенность, так как есть совпадение по фамилии и дате
            max_confidence = max(max_confidence, 0.85)
            logger.info(f"Найдено подтверждение по фамилии и дате рождения в ответе по email")

    logger.info(f"Результаты анализа второго этапа: {len(phones)} телефонов, уверенность {max_confidence:.2f}")
    return phones, max_confidence


def extract_phones_from_api_response(response, target_vk_id=None):
    """
    Улучшенное извлечение телефонных номеров из ответа API с более агрессивным поиском
    и поддержкой приоритетных баз данных.

    Args:
        response: Ответ API
        target_vk_id: Опциональный VK ID

    Returns:
        List[str]: Список найденных телефонных номеров
    """
    try:
        logger.info("Начинаем улучшенное извлечение телефонов из ответа API")

        # УЛУЧШЕНИЕ 1: Расширенный список приоритетных баз данных
        SOURCE_PRIORITY = {
            "Gosuslugi 2024": 15,
            "BolshayaPeremena": 14,
            "AlfaBank 2023 v2": 13,
            "ScanTour.ru": 12,
            "Resh.Edu": 11,
            "ProPostuplenie.ru": 10,
            "TrudVsem.ru": 9,
            "Dobro.ru": 8,
            "CDEK": 7,
            "Whoosh-bike": 6,
            "DNS (2022)": 5,
            "SushiMaster.ru": 5,
            "BurgerKing.ru": 5,
            "GloriaJeans": 5,
            "LeaderID": 4,
            "Book24": 4,
            "Rendez-Vous": 4,
            "Zoloto585.ru": 4,
            "Tokyo-city.ru": 3,
            "Pikabu": 3,
            "Metro-cc.ru": 3,
            "Adengi.ru": 3,
            "Mira1.ru": 3,
            "Orteka.ru": 2,
            "Oriflame.ru": 2,
            "Artek": 2
        }

        # УЛУЧШЕНИЕ 2: Проверка текстового представления на наличие пометки "Номер который нужно забирать"
        response_str = str(response)
        marked_phones = []

        if "Номер который нужно забирать" in response_str:
            logger.info("Найдена пометка 'Номер который нужно забирать'")

            # Разные варианты паттернов для номеров с такой пометкой
            marked_patterns = [
                r'📞Телефон:\s*(\d+)[^)]*Номер который нужно забирать',
                r'Телефон:\s*(\d+)[^)]*Номер который нужно забирать',
                r'телефон:\s*(\d+)[^)]*Номер который нужно забирать',
                r'\b(79\d{9})\b[^)]*Номер который нужно забирать'
            ]

            for pattern in marked_patterns:
                matches = re.findall(pattern, response_str)
                for match in matches:
                    digits = ''.join(c for c in match if c.isdigit())
                    if digits.startswith('79') and len(digits) == 11 and digits not in marked_phones:
                        marked_phones.append(digits)
                        logger.info(f"Найден приоритетный телефон с пометкой: {digits}")

            # Если есть телефоны с пометкой, возвращаем их немедленно
            if marked_phones:
                logger.info(f"Возвращаем {len(marked_phones)} приоритетных телефонов с пометкой")
                return marked_phones

        # УЛУЧШЕНИЕ 3: Обработка нестандартных типов ответа
        if not isinstance(response, dict):
            logger.warning(f"Неожиданный тип ответа API: {type(response)}. Ожидался dict.")
            # Поиск телефонов с помощью регулярных выражений в строковом представлении
            if response is not None:
                phone_patterns = [
                    r'📞Телефон:\s*(\d+)',
                    r'Телефон:\s*(\d+)',
                    r'телефон:\s*(\d+)',
                    r'Phone:\s*(\d+)',
                    r'[\'"]phone[\'"]:\s*[\'"](\d+)[\'"]',
                    r'[\'"]телефон[\'"]:\s*[\'"](\d+)[\'"]',
                    r'\b(79\d{9})\b',  # Формат 79XXXXXXXXX
                    r'\b(89\d{9})\b'  # Формат 89XXXXXXXXX
                ]

                found_phones = []
                for pattern in phone_patterns:
                    matches = re.findall(pattern, response_str)
                    for match in matches:
                        digits = ''.join(c for c in match if c.isdigit())
                        if (digits.startswith('79') or digits.startswith('89')) and len(digits) == 11:
                            # Нормализация номера (замена 8 на 7 в начале)
                            if digits.startswith('8'):
                                digits = '7' + digits[1:]
                            if digits not in found_phones:
                                found_phones.append(digits)

                return found_phones
            return []

        # Проверка структуры ответа
        if "List" not in response:
            logger.warning("Ключ 'List' отсутствует в ответе API")
            return extract_phones_recursive(response, target_vk_id)

        if not isinstance(response["List"], dict):
            logger.warning(f"Неожиданный тип для response['List']: {type(response['List'])}. Ожидался dict.")
            return extract_phones_recursive(response, target_vk_id)

        # УЛУЧШЕНИЕ 4: Более структурированный поиск телефонов по базам с учетом приоритетов
        all_phones = []
        phones_by_source = {}

        # Обрабатываем каждую базу данных
        for db_name, db_info in response["List"].items():
            # Пропускаем "No results found"
            if db_name == "No results found" or not isinstance(db_info, dict):
                continue

            # Проверка наличия Data
            if "Data" not in db_info or not isinstance(db_info["Data"], list):
                continue

            # УЛУЧШЕНИЕ 5: Проверка на наличие пометки в информации о базе
            source_info_str = str(db_info)
            if "Номер который нужно забирать" in source_info_str:
                logger.info(f"Найдена пометка 'Номер который нужно забирать' в базе {db_name}")
                marked_patterns = [
                    r'📞Телефон:\s*(\d+)[^)]*Номер который нужно забирать',
                    r'Телефон:\s*(\d+)[^)]*Номер который нужно забирать',
                    r'телефон:\s*(\d+)[^)]*Номер который нужно забирать'
                ]

                for pattern in marked_patterns:
                    matches = re.findall(pattern, source_info_str)
                    for match in matches:
                        digits = ''.join(c for c in match if c.isdigit())
                        if digits.startswith('79') and len(digits) == 11 and digits not in marked_phones:
                            marked_phones.append(digits)
                            logger.info(f"Найден приоритетный телефон с пометкой в базе {db_name}: {digits}")

                # Если есть телефоны с пометкой, возвращаем их
                if marked_phones:
                    logger.info(f"Возвращаем {len(marked_phones)} приоритетных телефонов с пометкой из базы {db_name}")
                    return marked_phones

            # УЛУЧШЕНИЕ 6: Прямой поиск телефонов в записях базы
            try:
                logger.debug(f"Обработка базы данных: {db_name}")

                source_phones = []
                for record in db_info["Data"]:
                    if not isinstance(record, dict):
                        continue

                    # Ищем поля с телефонами
                    for field_name, field_value in record.items():
                        if not field_value:
                            continue

                        field_lower = field_name.lower()
                        if "phone" in field_lower or "телефон" in field_lower or "тел" in field_lower:
                            # Извлекаем только цифры
                            digits = ''.join(c for c in str(field_value) if c.isdigit())

                            # Проверяем формат (должен начинаться с 79 и иметь длину 11 символов)
                            if (digits.startswith('79') or digits.startswith('89')) and len(digits) == 11:
                                # Нормализация номера (замена 8 на 7 в начале)
                                if digits.startswith('8'):
                                    digits = '7' + digits[1:]

                                # Проверяем, не содержит ли значение пометку
                                if "Номер который нужно забирать" in str(field_value):
                                    logger.info(f"Найден приоритетный телефон с пометкой в поле {field_name}: {digits}")
                                    return [digits]  # Сразу возвращаем приоритетный телефон

                                # Добавляем в список телефонов для текущей базы
                                if digits not in source_phones:
                                    source_phones.append(digits)

                # Если нашли телефоны, сохраняем их
                if source_phones:
                    phones_by_source[db_name] = source_phones
                    logger.info(f"Найдено {len(source_phones)} телефонов в базе {db_name}")
            except Exception as e:
                logger.error(f"Ошибка при обработке базы {db_name}: {e}")
                logger.error(traceback.format_exc())

        # УЛУЧШЕНИЕ 7: Если не нашли телефоны в полях напрямую, используем рекурсивный метод
        if not phones_by_source:
            recursive_phones = extract_phones_recursive(response, target_vk_id)
            if recursive_phones:
                logger.info(f"Найдено {len(recursive_phones)} телефонов рекурсивным методом")
                return recursive_phones

        # Сортируем источники по приоритету
        sorted_sources = sorted(
            phones_by_source.keys(),
            key=lambda src: SOURCE_PRIORITY.get(src, 0),
            reverse=True
        )

        # Собираем телефоны, начиная с наиболее приоритетных источников
        for source in sorted_sources:
            priority = SOURCE_PRIORITY.get(source, 0)
            logger.info(f"Обработка источника {source} с приоритетом {priority}")
            all_phones.extend(phones_by_source[source])

        # Удаляем дубликаты, сохраняя порядок (сначала из приоритетных источников)
        unique_phones = []
        for phone in all_phones:
            if phone not in unique_phones:
                unique_phones.append(phone)

        logger.info(f"Всего найдено уникальных телефонов: {len(unique_phones)}")
        return unique_phones

    except Exception as e:
        # Обрабатываем любые непредвиденные ошибки
        logger.error(f"Непредвиденная ошибка при извлечении телефонов: {e}")
        logger.error(traceback.format_exc())
        return []


def evaluate_phone_confidence(phone_entry, query_data):
    """
    Оценивает уверенность в правильности номера телефона (0.0-1.0)
    """
    confidence = 0.5  # Базовое значение

    # Фактор 1: Приоритет базы (до +0.3)
    if "priority" in phone_entry:
        if phone_entry["priority"] >= 8:
            confidence += 0.3
        elif phone_entry["priority"] >= 5:
            confidence += 0.2
        elif phone_entry["priority"] >= 3:
            confidence += 0.1

    # Фактор 2: Проверка совпадения ФИО (до +0.3)
    name_match = 0.0

    # Проверяем наличие ключа "record"
    if "record" in phone_entry:
        record = phone_entry["record"]

        for field, value in record.items():
            field_lower = field.lower()
            # Проверка фамилии
            if "фамилия" in field_lower or "lastname" in field_lower:
                if query_data["surname"].lower() in str(value).lower():
                    name_match += 0.15
            # Проверка имени
            if "имя" in field_lower or "firstname" in field_lower:
                if query_data["name"].lower() in str(value).lower():
                    name_match += 0.15

        confidence += name_match

        # Фактор 3: Проверка даты рождения (до +0.2)
        for field, value in record.items():
            field_lower = field.lower()
            if "рождения" in field_lower or "birth" in field_lower or "дата" in field_lower:
                # Проверка точного или частичного совпадения даты
                if query_data["birth_date"] in str(value) or query_data["birth_date"].replace(".", "-") in str(value):
                    confidence += 0.2
                    break

    # Фактор 4: Совпадение с другими найденными номерами (+0.2)
    if phone_entry.get("confirmed_count", 0) > 0:
        confidence += 0.2

    return min(confidence, 1.0)  # Макс. значение 1.0


def extract_emails_from_response(response):
    """
    Извлекает email адреса из ответа API
    """
    emails = []

    # Обработка структуры ответа
    if "List" in response and isinstance(response["List"], dict):
        for db_name, db_info in response["List"].items():
            if db_name == "No results found" or "Data" not in db_info:
                continue

            for record in db_info["Data"]:
                if not isinstance(record, dict):
                    continue

                for field_name, field_value in record.items():
                    field_lower = field_name.lower()
                    if "email" in field_lower or "почта" in field_lower or "mail" in field_lower:
                        if field_value and isinstance(field_value, str) and '@' in field_value:
                            email = field_value.lower().strip()
                            if email not in emails:
                                emails.append(email)

    # Сортировка по популярным доменам - приоритет gmail.com, mail.ru и т.д.
    domain_priority = {
        "gmail.com": 10,
        "mail.ru": 9,
        "yandex.ru": 8,
        "bk.ru": 7,
        "inbox.ru": 6
    }

    emails.sort(key=lambda e: domain_priority.get(e.split('@')[1], 0) if '@' in e and len(e.split('@')) > 1 else 0,
                reverse=True)

    return emails






